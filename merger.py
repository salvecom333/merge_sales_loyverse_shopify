import pandas as pd
import os
import tkinter.filedialog as fd
import re
from openpyxl import load_workbook
from tkinter import simpledialog
import tkinter
import numpy
import matplotlib.pyplot as plt
# change the slash from “\” to “/”, if you are using Windows devices

root = fd.Tk()
root.withdraw() #use to hide tkinter window

currdir = os.getcwd()
input_file_path = fd.askdirectory(parent=root, initialdir=currdir, title='Please select INPUT folder')
if len(input_file_path) > 0:
    print ("You chose %s" % input_file_path)

output_file_path = fd.askdirectory(parent=root, initialdir=currdir, title='Please select OUTPUT folder')
if len(output_file_path) > 0:
    print ("You chose %s" % output_file_path)


def cafe_collect(collection):
    collection['Category'] = collection['Category'].replace({'Cold Coffee Products': 'Cold Coffee Products(Cafe)',
                                               'Cafesito icecream':'Cafesito icecream(Cafe)',
                                               'CFE bills':'CFE bills(Cafe)',
                                               'Extra shots':'Extra shots(Cafe)',
                                               'Cold Coffee Products': 'Cold Coffee Products(Cafe)',
                                               'Hot Coffee Products':'Hot Coffee Products(Cafe)',
                                               'Loose Leaf Teas':'Loose Leaf Teas(Cafe)',
                                               'Refillable Spices/accessories':'Refillable Spices/accessories(Cafe)',
                                               'Whole bean coffee bags':'Whole bean coffee bags(Cafe)',
                                               'Accessories(Online)':'Accessories(Online)(Cafe)',
                                               'Calimax service(Online)':'Calimax service(Online)(Cafe)',
                                               'Cold(Online)':'Cold(Online)(Cafe)',
                                               'Extra shots(Online)':'Extra shots(Online)(Cafe)',
                                               'Home page(Online)':'Home page(Online)(Cafe)',
                                               'Hot(Online)':'Hot(Online)(Cafe)',
                                               'Loose leaf teas(Online)':'Loose leaf teas(Online)(Cafe)',
                                               'Pacific Electric Coffee Co Merch(Online)':'Pacific Electric Coffee Co Merch(Online)(Cafe)',
                                               'Whole bean coffee(Online)':'Whole bean coffee(Online)(Cafe)',
                                               'Nicaragua Coffee(Online)':'Nicaragua Coffee(Online)(Cafe)'})

def serafina_collect(collection):
    collection['Category'] = collection['Category'].replace({'Nana products(Online)': 'Nana products(Online)(Serafina)',
                                               'Serafina cheese':'Serafina cheese(Serafina)',
                                               'Serafina cheese(Online)':'Serafina cheese(Online)(Serafina)',
                                               'Mom baked goods':'Mom baked goods(Serafina)'})

def keto_collect(collection):
    collection['Category'] = collection['Category'].replace({'Mary Baked Goods': 'Mary Baked Goods(Keto)',
                                               'Mary keto baked goods(Online)':'Mary keto baked goods(Online)(Keto)'})

def org_items(data):
    data["Variant"][data['Product'].str.contains("Home", case=False)] = "Whole bean coffee bags"
    data["Variant"][data['Product'].str.contains("Cold", case=False)] = "Cold Coffee Products"
    data["Variant"][data['Product'].str.contains("victor", case=False)] = "Victor Art Sales"
    data["Variant"][data['Product'].str.contains("roast", case=False)] = "Whole bean coffee bags"
    data["Variant"][data['Product'].str.contains("hot", case=False)] = "Hot Coffee Products"
    data["Variant"][data['Product'].str.contains("stray", case=False)] = "Linda Stray Dog Icecream"
    data["Variant"][data['Product'].str.contains("chism", case=False)] = "Chism baked goods"
    data["Variant"][data['Product'].str.contains("blaine", case=False)] = "Blaine baked goods"
    data["Variant"][data['Product'].str.contains("claudia", case=False)] = "Claudia Baked Goods"
    data["Variant"][data['Product'].str.contains("extra", case=False)] = "Extra shots"
    data["Variant"][data['Product'].str.contains("gabriela", case=False)] = "Gabriela Soto(Vasos)"
    data["Variant"][data['Product'].str.contains("gloria", case=False)] = "Gloria Products"
    data["Variant"][data['Product'].str.contains("loose", case=False)] = "Loose Leaf Teas"
    data["Variant"][data['Product'].str.contains("CFE", case=False)] = "CFE bills"
    data["Variant"][data['Product'].str.contains("Mary keto", case=False)] = "Mary Baked Goods"
    data["Variant"][data['Product'].str.contains("Mary Lou", case=False)] = "Mary Lou pies"
    data["Variant"][data['Product'].str.contains("Accessories", case=False)] = "Refillable Spices/accessories"
    data["Variant"][data['Product'].str.contains("Rachel", case=False)] = "Rachel Italian Button"
    data["Variant"][data['Product'].str.contains("Vicky", case=False)] = "Vicky Baked Goods"
    data["Variant"][data['Product'].str.contains("Calimax", case=False)] = "Calimax service"
    data["Variant"][data['Product'].str.contains("Nana", case=False)] = "Mom baked goods"
    data["Variant"][data['Product'].str.contains("Serafina", case=False)] = "Serafina cheese"
    data["Variant"][data['Product'].str.contains("Nancy", case=False)] = "Nancy baked goods"
    data["Variant"][data['Product'].str.contains("Chris", case=False)] = "Chris Corral baked goods"
    data["Variant"][data['Product'].str.contains("cava", case=False)] = "C & G cava sales"
    data["Variant"][data['Product'].str.contains("Arlene", case=False)] = "Arlene Dentist"
    data["Variant"][data['Product'].str.contains("Diana", case=False)] = "Diana baked goods"
    data.rename(columns = {'Variant':'Category'}, inplace = True)
    data.rename(columns = {'Product':'Item name'}, inplace = True)
    data.rename(columns = {'Net Units Sold':'Items sold'}, inplace = True)
    data.rename(columns = {'Net Sales ($)':'Net sales'}, inplace = True)
    return(data)
    #for ele in data['Product']:
        #print(ele)
        #if 'Cold' in ele:
            #data['Variant'].replace({'':'Cold Coffee Products'})

def add_title(frame):
    title = 'Outputs'
    header = 'Items by Clients below:vvvv'

   # Add MultiIndex Header
    frame.columns = pd.MultiIndex.from_product([[header], frame.columns])
    styler = frame.style.set_caption(title).set_table_styles([{
        'selector': 'caption',
        'props': [
            ('color', 'red'),
            ('font-size', '15px'),
            ('font-style', 'italic'),
            ('font-weight', 'bold'),
            ('text-align', 'center')
        ]
    }])
    return(frame)

def org_item_2(collection):

    Cafe_sheet = collection['Category'].str.contains('Cafe')
    cafesito = collection[Cafe_sheet]
    cafesito = calculation(cafesito)

    Blaine_sheet = collection['Category'].str.contains('Blaine', case=False)
    Blaine = collection[Blaine_sheet]
    Blaine = calculation(Blaine)

    Chism_sheet = collection['Category'].str.contains('Chism', case=False)
    Chism = collection[Chism_sheet]
    Chism = calculation(Chism)

    Claudia_sheet = collection['Category'].str.contains('Claudia', case=False)
    Claudia = collection[Claudia_sheet]
    Claudia = calculation(Claudia)

    Gabriela_sheet = collection['Category'].str.contains('Gabriela', case=False)
    Gabriela = collection[Gabriela_sheet]
    Gabriela = calculation(Gabriela)

    Gloria_sheet = collection['Category'].str.contains('Gloria', case=False)
    Gloria = collection[Gloria_sheet]
    Gloria = calculation(Gloria)

    Linda_sheet = collection['Category'].str.contains('Icecream', case=False)
    Linda = collection[Linda_sheet]
    Linda = calculation(Linda)

    Maryketo_sheet = collection['Category'].str.contains('Keto', case=False)
    Maryketo = collection[Maryketo_sheet]
    Maryketo = calculation(Maryketo)

    MaryLou_sheet = collection['Category'].str.contains('Mary Lou', case=False)
    MaryLou = collection[MaryLou_sheet]
    MaryLou = calculation(MaryLou)

    Mom_sheet = collection['Category'].str.contains('Serafina', case=False)
    Mom = collection[Mom_sheet]
    Mom = calculation(Mom)

    Vicky_sheet = collection['Category'].str.contains('Vicky', case=False)
    Vicky = collection[Vicky_sheet]
    Vicky = calculation(Vicky)

    Victor_sheet = collection['Category'].str.contains('Victor', case=False)
    Victor = collection[Victor_sheet]
    Victor = calculation(Victor)

    Nancy_sheet = collection['Category'].str.contains('Nancy', case=False)
    Nancy = collection[Nancy_sheet]
    Nancy = calculation(Nancy)

    Chris_sheet = collection['Category'].str.contains('Chris', case=False)
    Chris = collection[Chris_sheet]
    Chris = calculation(Chris)

    Cava_sheet = collection['Category'].str.contains('cava', case=False)
    Cava = collection[Cava_sheet]
    Cava = calculation(Cava)

    Arlene_sheet = collection['Category'].str.contains('Arlene', case=False)
    Arlene = collection[Arlene_sheet]
    Arlene = calculation(Arlene)

    Diana_sheet = collection['Category'].str.contains('Diana', case=False)
    Diana = collection[Diana_sheet]
    Diana = calculation(Diana)

    add_title(cafesito)
    add_title(Blaine)
    add_title(Chism)
    add_title(Claudia)
    add_title(Gabriela)
    add_title(Gloria)
    add_title(Linda)
    add_title(Maryketo)
    add_title(MaryLou)
    add_title(Mom)
    add_title(Vicky)
    add_title(Victor)
    add_title(Nancy)
    add_title(Chris)
    add_title(Cava)
    add_title(Arlene)
    add_title(Diana)

    cafesito.to_excel(writer, sheet_name = 'Cafesito Lindo', startrow=writer.sheets['Cafesito Lindo'].max_row, index = True,header= "Items of Clients below:vvvv")
    Blaine.to_excel(writer, sheet_name = 'Blaine_reciept', startrow=writer.sheets['Blaine_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Chism.to_excel(writer, sheet_name = 'Chism_reciept', startrow=writer.sheets['Chism_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Claudia.to_excel(writer, sheet_name = 'Claudia_reciept', startrow=writer.sheets['Claudia_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Gabriela.to_excel(writer, sheet_name = 'Gabriela_reciept', startrow=writer.sheets['Gabriela_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Gloria.to_excel(writer, sheet_name = 'Gloria_reciept', startrow=writer.sheets['Gloria_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Linda.to_excel(writer, sheet_name = 'Linda_reciept', startrow=writer.sheets['Linda_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Maryketo.to_excel(writer, sheet_name = 'Maryketo_reciept', startrow=writer.sheets['Maryketo_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    MaryLou.to_excel(writer, sheet_name = 'MaryLou_reciept', startrow=writer.sheets['MaryLou_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Mom.to_excel(writer, sheet_name = 'Mom_reciept', startrow=writer.sheets['Mom_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Vicky.to_excel(writer, sheet_name = 'Vicky_reciept', startrow=writer.sheets['Vicky_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Victor.to_excel(writer, sheet_name = 'Victor_reciept', startrow=writer.sheets['Victor_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Nancy.to_excel(writer, sheet_name = 'Nancy_reciept', startrow=writer.sheets['Nancy_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Chris.to_excel(writer, sheet_name = 'Chris_reciept', startrow=writer.sheets['Chris_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Cava.to_excel(writer, sheet_name = 'Cava_reciept', startrow=writer.sheets['Cava_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Arlene.to_excel(writer, sheet_name = 'Arlene_reciept', startrow=writer.sheets['Arlene_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    Diana.to_excel(writer, sheet_name = 'Diana_reciept', startrow=writer.sheets['Diana_reciept'].max_row, index = True,header= "Items of Clients below:vvvv")
    ##organize(frame)
def calculation(category):
    #Cafe_sheet = category['Category'].str.contains('Cafe')
    #cafesito = category[Cafe_sheet]

    item_sum = category['Items sold'].sum(axis = 0)
    sale_sum = category['Net sales'].sum(axis = 0)
    cat_total = pd.DataFrame({'Category':['Total:'], 'Items sold':[item_sum], 'Net sales':[sale_sum]})
    category = pd.concat([category, cat_total])


    if category['Category'].str.contains('Cafe').any() == True:
        consignment_fee = pd.DataFrame({'Category':['Consignment fee->'], 'Net sales':['No consignment fee']})
        category = pd.concat([category, consignment_fee])
        con_sum = 0
    elif category['Category'].str.contains('Mary baked goods', case = False).any() == True:
        con_sum = sale_sum * 0.20 * -1
        consignment_fee = pd.DataFrame({'Category':['Consignment fee->'],'Items sold':['{:.1%}'.format(0.20)], 'Net sales':[con_sum]})
        category = pd.concat([category, consignment_fee])
    else:
        con_sum = sale_sum * 0.16 * -1
        consignment_fee = pd.DataFrame({'Category':['Consignment fee->'],'Items sold':['{:.1%}'.format(0.16)], 'Net sales':[con_sum]})
        category = pd.concat([category, consignment_fee])

    total_sum = sale_sum + con_sum
    net_total = pd.DataFrame({'Items sold':['Net Total='], 'Net sales':['${}'.format(total_sum)]})
    category = pd.concat([category, net_total])

    peso_sum = total_sum * 19
    peso_sum = round(peso_sum, 2)
    convert_peso = pd.DataFrame({'Items sold':['Peso Total='], 'Net sales':['{}p'.format(peso_sum)]})
    category = pd.concat([category, convert_peso])
    return(category)
        #category.to_excel(writer, sheet_name = 'Cafesito Lindo')


def organize(collection):
    Cafe_sheet = collection['Category'].str.contains('Cafe')
    cafesito = collection[Cafe_sheet]
    cafesito = calculation(cafesito)

    Blaine_sheet = collection['Category'].str.contains('Blaine', case=False)
    Blaine = collection[Blaine_sheet]
    Blaine = calculation(Blaine)

    Chism_sheet = collection['Category'].str.contains('Chism', case=False)
    Chism = collection[Chism_sheet]
    Chism = calculation(Chism)

    Claudia_sheet = collection['Category'].str.contains('Claudia', case=False)
    Claudia = collection[Claudia_sheet]
    Claudia = calculation(Claudia)

    Gabriela_sheet = collection['Category'].str.contains('Gabriela', case=False)
    Gabriela = collection[Gabriela_sheet]
    Gabriela = calculation(Gabriela)

    Gloria_sheet = collection['Category'].str.contains('Gloria', case=False)
    Gloria = collection[Gloria_sheet]
    Gloria = calculation(Gloria)

    Linda_sheet = collection['Category'].str.contains('Icecream', case=False)
    Linda = collection[Linda_sheet]
    Linda = calculation(Linda)

    Maryketo_sheet = collection['Category'].str.contains('Keto', case=False)
    Maryketo = collection[Maryketo_sheet]
    Maryketo = calculation(Maryketo)

    MaryLou_sheet = collection['Category'].str.contains('Mary Lou', case=False)
    MaryLou = collection[MaryLou_sheet]
    MaryLou = calculation(MaryLou)

    Mom_sheet = collection['Category'].str.contains('Serafina', case=False)
    Mom = collection[Mom_sheet]
    Mom = calculation(Mom)

    Vicky_sheet = collection['Category'].str.contains('Vicky', case=False)
    Vicky = collection[Vicky_sheet]
    Vicky = calculation(Vicky)

    Victor_sheet = collection['Category'].str.contains('Victor', case=False)
    Victor = collection[Victor_sheet]
    Victor = calculation(Victor)

    Nancy_sheet = collection['Category'].str.contains('Nancy', case=False)
    Nancy = collection[Nancy_sheet]
    Nancy = calculation(Nancy)

    Chris_sheet = collection['Category'].str.contains('Chris', case=False)
    Chris = collection[Chris_sheet]
    Chris = calculation(Chris)

    Cava_sheet = collection['Category'].str.contains('cava', case=False)
    Cava = collection[Cava_sheet]
    Cava = calculation(Cava)

    Arlene_sheet = collection['Category'].str.contains('Arlene', case=False)
    Arlene = collection[Arlene_sheet]
    Arlene = calculation(Arlene)

    Diana_sheet = collection['Category'].str.contains('Diana', case=False)
    Diana = collection[Diana_sheet]
    Diana = calculation(Diana)

    cafesito.to_excel(writer, sheet_name = 'Cafesito Lindo')
    Blaine.to_excel(writer, sheet_name = 'Blaine_reciept')
    Chism.to_excel(writer, sheet_name = 'Chism_reciept')
    Claudia.to_excel(writer, sheet_name = 'Claudia_reciept')
    Gabriela.to_excel(writer, sheet_name = 'Gabriela_reciept')
    Gloria.to_excel(writer, sheet_name = 'Gloria_reciept')
    Linda.to_excel(writer, sheet_name = 'Linda_reciept')
    Maryketo.to_excel(writer, sheet_name = 'Maryketo_reciept')
    MaryLou.to_excel(writer, sheet_name = 'MaryLou_reciept')
    Mom.to_excel(writer, sheet_name = 'Mom_reciept')
    Vicky.to_excel(writer, sheet_name = 'Vicky_reciept')
    Victor.to_excel(writer, sheet_name = 'Victor_reciept')
    Nancy.to_excel(writer, sheet_name = 'Nancy_reciept')
    Chris.to_excel(writer, sheet_name = 'Chris_reciept')
    Cava.to_excel(writer, sheet_name = 'Cava_reciept')
    Arlene.to_excel(writer, sheet_name = 'Arlene_reciept')
    Diana.to_excel(writer, sheet_name = 'Diana_reciept')


    #writer.close()

#create a list to store all the file references of the input folder using the listdir function from the os library.
#To see the contents of a library (like the listdir function, you can use the dir function on the library name).
#Use dir(library_name) to list contents

excel_file_list = os.listdir(input_file_path)
#seperate item csvs with the net csvs***
net_list = [file for file in excel_file_list if 'category' in file or 'Collection' in file]
item_list = [file for file in excel_file_list if 'item'in file or 'Product' in file]
##itemdf2.groupby(itemdf2.columns, axis=1).agg(numpy.max)
#itemdf2.to_excel(output_file_path+file_name)
##print(itemdf4.to_string())
###itemdf2.rename(columns = {'Variant':'Category'}, inplace = True)
#print(net_list)
##print(item_list)
##print(excel_file_list)

itemdf = pd.DataFrame()
for file in item_list:
    if file.endswith(".csv"):
        itemdf1 = pd.read_csv(file)
        itemdf = pd.concat([itemdf, pd.DataFrame.from_records(itemdf1)])

itemdf2 = itemdf[['Category','Item name', 'Items sold', 'Net sales', 'Product', 'Variant', 'Net Units Sold', 'Net Sales ($)']]
itemdf2 = itemdf2.fillna('')
itemdf2['Product'] = itemdf2['Product'].astype(str) +'('+ itemdf2.astype(str)['Variant'] +')'
itemdf2['Variant'].iloc[:] = ''
itemdf2['Product'] = itemdf2['Product'].str.removesuffix("()")
itemdf2['Product'] = itemdf2['Product'] + '(Online)'
itemdf2['Product'] = itemdf2['Product'].replace({'(Online)': ''})
org_items(itemdf2)
itemdf3 = itemdf2.replace(r'^\s*$', numpy.nan, regex=True)
itemdf4 = itemdf3.groupby(lambda x:x, axis=1).sum()
itemdf4 = itemdf4.loc[itemdf4["Category"] != 0]
cafe_collect(itemdf4)
serafina_collect(itemdf4)
keto_collect(itemdf4)
##print(itemdf4.to_string())
###organize(itemdf4)
#Once each file opens, use the append function to start consolidating the data stored in multiple files

#create a new, blank dataframe, to handle the excel file imports
df = pd.DataFrame()

#Run a for loop to loop through each file in the list
for file in net_list:
   #check for .csv suffix files only
   if file.endswith(".csv"):
       #create a new dataframe to read/open each Excel file from the list of files created above
       df1 = pd.read_csv(file)
       #append each file into the original empty dataframe
       df = pd.concat([df, pd.DataFrame.from_records(df1)])

#transfer final output to an Excel (xlsx) file on the output path

df2 = df[['Category', 'Items sold', 'Net sales', 'Collection', 'Net Units Sold', 'Net Sales ($)']]
df2['Collection'] = df2['Collection'] + '(Online)'
df2.rename(columns = {'Collection':'Category'}, inplace = True)
df2.rename(columns = {'Net Units Sold':'Items sold'}, inplace = True)
df2.rename(columns = {'Net Sales ($)':'Net sales'}, inplace = True)


#df2['Category'] = df2['Category'].replace({'Cold Coffee Products': 'Cold Coffee Products(Cafe)'})
cafe_collect(df2)
serafina_collect(df2)
keto_collect(df2)
#organize(df2)
df3 = df2.groupby(lambda x:x, axis=1).sum()
    ##print(df3.to_string())
date = simpledialog.askstring(title="Reciepts",
                              prompt="Type dates for reciepts:\n(MM-DD-YY-MM-DD-YY)format")

#date = input('Type dates for reciepts:\n(MM-DD-YY-MM-DD-YY)format\n')

load_file_path = "{}/reciepts{}.xlsx".format(output_file_path, date)
file_name = "reciepts{}.xlsx".format(date)
#itemdf2.to_excel(output_file_path+file_name)
#print(file_name)
#print(load_file_path)
#df3.to_excel(output_file_path+file_name)


with pd.ExcelWriter(load_file_path, engine = 'openpyxl') as writer:
    df3.to_excel(writer, sheet_name = 'All_Collections')

    itemdf4.to_excel(writer, sheet_name='All_Collections', startrow=writer.sheets['All_Collections'].max_row, index = True,header= "Items of Clients below:vvvv")
    ##itemdf4.to_excel(writer, sheet_name = 'All_Collections')
    ##df3.to_excel(writer, sheet_name = 'All_Collections')
    ##organize(itemdf4)
    ###Comment organize(df3) if you only want items function to work and viceversa
    organize(df3)

    org_item_2(itemdf4)


    #org_item_2(itemdf4)
